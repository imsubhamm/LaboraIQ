import ipaddress
import socket
import time
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Annotated, Any, TypeVar, cast

from fastapi import (
    APIRouter,
    Depends,
    File,
    HTTPException,
    Query,
    Request,
    Response,
    UploadFile,
    status,
)
from fastapi.encoders import jsonable_encoder
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.analyzer_orders import create_queued_attempt, process_queued_orders
from app.audit import record_event
from app.auth import AuthContext, require_permission
from app.config import get_settings
from app.database import get_db
from app.models import (
    Analyzer,
    AnalyzerConnectionEvent,
    AnalyzerOrderAttempt,
    AnalyzerParameterMapping,
    AnalyzerTestMapping,
    AnalyzerWorklistItem,
    AuditEvent,
    Branch,
    Department,
    Invoice,
    LabOrder,
    OrderTest,
    Organization,
    Patient,
    PatientHistory,
    Permission,
    Role,
    Specimen,
    Status,
    TestCatalogItem,
    TestCatalogParameter,
    User,
    UserRoleAssignment,
)
from app.schemas import (
    AnalyzerConnectionEventRead,
    AnalyzerConnectionTestRead,
    AnalyzerCreate,
    AnalyzerMappingStatusUpdate,
    AnalyzerOrderAttemptRead,
    AnalyzerOrderProcessRead,
    AnalyzerRead,
    AnalyzerTestMappingCreate,
    AnalyzerTestMappingRead,
    AnalyzerUpdate,
    AnalyzerWorklistCancel,
    AnalyzerWorklistRead,
    AssignmentCreate,
    AssignmentRead,
    AuditEventRead,
    BranchCreate,
    BranchRead,
    BranchUpdate,
    DepartmentCreate,
    DepartmentRead,
    DepartmentUpdate,
    IntakeCreate,
    IntakeRead,
    OrganizationCreate,
    OrganizationRead,
    OrganizationUpdate,
    Page,
    PatientLookupRead,
    PaymentCreate,
    PaymentRead,
    PaymentSummary,
    PermissionRead,
    RoleCreate,
    RoleRead,
    SpecimenCollect,
    SpecimenDecision,
    SpecimenRead,
    SpecimenWorkflowRead,
    TestCatalogRead,
    TestMasterCreate,
    TestMasterImportRead,
    TestMasterRead,
    TestParameterCreate,
    TestParameterRead,
    TestParameterUpdate,
    UserCreate,
    UserRead,
    UserUpdate,
)

router = APIRouter()
Db = Annotated[Session, Depends(get_db)]
ModelT = TypeVar("ModelT")


def approved_analyzer_overlay_targets() -> set[ipaddress.IPv4Address | ipaddress.IPv6Address]:
    targets: set[ipaddress.IPv4Address | ipaddress.IPv6Address] = set()
    for value in get_settings().analyzer_overlay_targets:
        try:
            targets.add(ipaddress.ip_address(value))
        except ValueError:
            continue
    return targets


def page(db: Session, statement: Any, model: type[ModelT], limit: int, offset: int) -> Page[ModelT]:
    total = db.scalar(select(func.count()).select_from(statement.subquery())) or 0
    items = list(db.scalars(statement.limit(limit).offset(offset)).all())
    return Page[model](items=items, total=total, limit=limit, offset=offset)  # type: ignore[valid-type]


def commit(db: Session) -> None:
    try:
        db.commit()
    except IntegrityError as error:
        db.rollback()
        raise HTTPException(
            status_code=409, detail="A record with this code already exists"
        ) from error


def flush(db: Session) -> None:
    try:
        db.flush()
    except IntegrityError as error:
        db.rollback()
        raise HTTPException(
            status_code=409, detail="A record with this code already exists"
        ) from error


def get_tenant_record(
    db: Session, model: type[ModelT], record_id: uuid.UUID, context: AuthContext
) -> ModelT:
    record = db.scalar(
        select(model).where(
            model.id == record_id,  # type: ignore[attr-defined]
            model.organization_id == context.organization_id,  # type: ignore[attr-defined]
        )
    )
    if record is None:
        raise HTTPException(status_code=404, detail="Record not found")
    return record


def snapshot(record: Any, keys: list[str]) -> dict[str, Any]:
    return cast(
        dict[str, Any],
        jsonable_encoder({key: getattr(record, key) for key in keys}),
    )


ANALYZER_FIELDS = [
    "branch_id",
    "code",
    "vendor",
    "model",
    "protocol",
    "host",
    "port",
    "connection_mode",
    "connection_timeout_seconds",
    "retry_limit",
    "heartbeat_interval_seconds",
    "connection_status",
    "status",
]


@router.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@router.get("/ready")
def ready(db: Db) -> dict[str, str]:
    db.execute(select(1))
    return {"status": "ready"}


@router.get("/analyzers", response_model=Page[AnalyzerRead])
def list_analyzers(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.read"))],
    branch_id: uuid.UUID | None = None,
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
) -> Page[AnalyzerRead]:
    statement = select(Analyzer).where(Analyzer.organization_id == context.organization_id)
    if not context.is_organization_scoped:
        statement = statement.where(Analyzer.branch_id.in_(context.branch_ids))
    if branch_id:
        if not context.can_access_branch(branch_id):
            raise HTTPException(status_code=403, detail="Branch access denied")
        statement = statement.where(Analyzer.branch_id == branch_id)
    return page(db, statement.order_by(Analyzer.code, Analyzer.id), AnalyzerRead, limit, offset)


@router.post("/analyzers", response_model=AnalyzerRead, status_code=201)
def create_analyzer(
    payload: AnalyzerCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.manage"))],
) -> Analyzer:
    get_branch(payload.branch_id, db, context)
    analyzer = Analyzer(
        **payload.model_dump(),
        organization_id=context.organization_id,
        created_by=context.user_id,
        updated_by=context.user_id,
    )
    db.add(analyzer)
    flush(db)
    record_event(
        db,
        request,
        context,
        event_type="analyzer.created",
        entity_type="analyzer",
        entity_id=analyzer.id,
        branch_id=analyzer.branch_id,
        action="create",
        new=snapshot(analyzer, ANALYZER_FIELDS),
    )
    commit(db)
    return analyzer


@router.patch("/analyzers/{analyzer_id}", response_model=AnalyzerRead)
def update_analyzer(
    analyzer_id: uuid.UUID,
    payload: AnalyzerUpdate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.manage"))],
) -> Analyzer:
    analyzer = get_tenant_record(db, Analyzer, analyzer_id, context)
    if not context.can_access_branch(analyzer.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    previous = snapshot(analyzer, ANALYZER_FIELDS)
    changes = payload.model_dump(exclude_unset=True)
    for key, value in changes.items():
        setattr(analyzer, key, value)
    if "host" in changes or "port" in changes:
        analyzer.connection_status = "never_tested"
        analyzer.last_connection_error = None
    analyzer.updated_by = context.user_id
    record_event(
        db,
        request,
        context,
        event_type="analyzer.updated",
        entity_type="analyzer",
        entity_id=analyzer.id,
        branch_id=analyzer.branch_id,
        action="update",
        previous=previous,
        new=snapshot(analyzer, ANALYZER_FIELDS),
    )
    commit(db)
    return analyzer


def get_accessible_analyzer(db: Session, analyzer_id: uuid.UUID, context: AuthContext) -> Analyzer:
    analyzer = get_tenant_record(db, Analyzer, analyzer_id, context)
    if not context.can_access_branch(analyzer.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    return analyzer


def validate_private_analyzer_target(host: str) -> None:
    address = ipaddress.ip_address(host)
    is_approved_private_target = (
        address.is_private or address in approved_analyzer_overlay_targets()
    )
    if (
        not is_approved_private_target
        or address.is_loopback
        or address.is_link_local
        or address.is_multicast
        or address.is_unspecified
    ):
        raise HTTPException(
            status_code=422,
            detail="Analyzer connection tests are restricted to private network addresses",
        )


def probe_analyzer_connection(
    analyzer: Analyzer,
    request: Request,
    db: Session,
    context: AuthContext,
    event_type: str,
) -> AnalyzerConnectionTestRead:
    validate_private_analyzer_target(analyzer.host)
    tested_at = datetime.now(UTC)
    analyzer.last_connection_test_at = tested_at
    final_success = False
    final_latency: int | None = None
    final_message = "Connection failed"
    attempts = analyzer.retry_limit + 1
    for attempt in range(1, attempts + 1):
        started = time.monotonic()
        try:
            with socket.create_connection(
                (analyzer.host, analyzer.port), timeout=analyzer.connection_timeout_seconds
            ):
                final_latency = max(0, round((time.monotonic() - started) * 1000))
                final_success = True
                final_message = "TCP connection established"
        except OSError as error:
            final_latency = max(0, round((time.monotonic() - started) * 1000))
            final_message = str(error)[:500] or error.__class__.__name__
        db.add(
            AnalyzerConnectionEvent(
                organization_id=context.organization_id,
                branch_id=analyzer.branch_id,
                analyzer_id=analyzer.id,
                event_type=event_type,
                attempt=attempt,
                success=final_success,
                latency_ms=final_latency,
                message=final_message,
                correlation_id=request.state.correlation_id,
                occurred_at=datetime.now(UTC),
            )
        )
        if final_success:
            attempts = attempt
            break
    analyzer.connection_status = "connected" if final_success else "error"
    analyzer.last_connection_error = None if final_success else final_message
    if final_success:
        analyzer.last_connected_at = tested_at
    record_event(
        db,
        request,
        context,
        event_type=f"analyzer.{event_type}",
        entity_type="analyzer",
        entity_id=analyzer.id,
        branch_id=analyzer.branch_id,
        action=event_type,
        new={
            "connection_status": analyzer.connection_status,
            "attempts": attempts,
            "latency_ms": final_latency,
            "message": final_message,
        },
    )
    commit(db)
    return AnalyzerConnectionTestRead(
        analyzer_id=analyzer.id,
        connection_status=analyzer.connection_status,
        attempts=attempts,
        success=final_success,
        latency_ms=final_latency,
        message=final_message,
        tested_at=tested_at,
    )


@router.post("/analyzers/{analyzer_id}/connection-test", response_model=AnalyzerConnectionTestRead)
def test_analyzer_connection(
    analyzer_id: uuid.UUID,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.manage"))],
) -> AnalyzerConnectionTestRead:
    analyzer = get_accessible_analyzer(db, analyzer_id, context)
    return probe_analyzer_connection(analyzer, request, db, context, "connection_test")


@router.post("/analyzers/{analyzer_id}/heartbeat", response_model=AnalyzerConnectionTestRead)
def heartbeat_analyzer(
    analyzer_id: uuid.UUID,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.manage"))],
) -> AnalyzerConnectionTestRead:
    analyzer = get_accessible_analyzer(db, analyzer_id, context)
    return probe_analyzer_connection(analyzer, request, db, context, "heartbeat")


@router.get(
    "/analyzers/{analyzer_id}/connection-events",
    response_model=list[AnalyzerConnectionEventRead],
)
def list_analyzer_connection_events(
    analyzer_id: uuid.UUID,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
) -> list[AnalyzerConnectionEvent]:
    analyzer = get_accessible_analyzer(db, analyzer_id, context)
    return list(
        db.scalars(
            select(AnalyzerConnectionEvent)
            .where(AnalyzerConnectionEvent.analyzer_id == analyzer.id)
            .order_by(AnalyzerConnectionEvent.occurred_at.desc())
            .limit(limit)
        ).all()
    )


def analyzer_mapping_read(db: Session, mapping: AnalyzerTestMapping) -> AnalyzerTestMappingRead:
    test = db.get(TestCatalogItem, mapping.test_id)
    parameter_ids = [item.parameter_id for item in mapping.parameters]
    catalog_parameters = {
        item.id: item
        for item in db.scalars(
            select(TestCatalogParameter).where(TestCatalogParameter.id.in_(parameter_ids))
        ).all()
    }
    return AnalyzerTestMappingRead(
        id=mapping.id,
        analyzer_id=mapping.analyzer_id,
        test_id=mapping.test_id,
        lis_test_code=test.code if test else "Unknown",
        test_name=test.name if test else "Unknown test",
        machine_test_code=mapping.machine_test_code,
        status=mapping.status,
        parameters=[
            {
                "id": item.id,
                "parameter_id": item.parameter_id,
                "parameter_name": catalog_parameters[item.parameter_id].name,
                "lis_parameter_code": catalog_parameters[item.parameter_id].external_code,
                "machine_parameter_code": item.machine_parameter_code,
                "unit": item.unit,
            }
            for item in mapping.parameters
            if item.parameter_id in catalog_parameters
        ],
        created_at=mapping.created_at,
        updated_at=mapping.updated_at,
    )


@router.get("/analyzers/{analyzer_id}/mappings", response_model=list[AnalyzerTestMappingRead])
def list_analyzer_mappings(
    analyzer_id: uuid.UUID,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.read"))],
) -> list[AnalyzerTestMappingRead]:
    analyzer = get_tenant_record(db, Analyzer, analyzer_id, context)
    if not context.can_access_branch(analyzer.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    mappings = db.scalars(
        select(AnalyzerTestMapping)
        .where(AnalyzerTestMapping.analyzer_id == analyzer.id)
        .order_by(AnalyzerTestMapping.machine_test_code, AnalyzerTestMapping.id)
    ).all()
    return [analyzer_mapping_read(db, item) for item in mappings]


@router.post(
    "/analyzers/{analyzer_id}/mappings",
    response_model=AnalyzerTestMappingRead,
    status_code=201,
)
def save_analyzer_mapping(
    analyzer_id: uuid.UUID,
    payload: AnalyzerTestMappingCreate,
    request: Request,
    response: Response,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.manage"))],
) -> AnalyzerTestMappingRead:
    analyzer = get_tenant_record(db, Analyzer, analyzer_id, context)
    if not context.can_access_branch(analyzer.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    test = get_tenant_record(db, TestCatalogItem, payload.test_id, context)
    supplied_parameter_ids = [item.parameter_id for item in payload.parameters]
    if len(supplied_parameter_ids) != len(set(supplied_parameter_ids)):
        raise HTTPException(status_code=422, detail="Each LIS parameter can be mapped only once")
    machine_parameter_codes = [
        item.machine_parameter_code.strip().upper() for item in payload.parameters
    ]
    if len(machine_parameter_codes) != len(set(machine_parameter_codes)):
        raise HTTPException(status_code=422, detail="Machine parameter codes must be unique")
    valid_parameter_ids = set(
        db.scalars(
            select(TestCatalogParameter.id).where(TestCatalogParameter.test_id == test.id)
        ).all()
    )
    if not set(supplied_parameter_ids).issubset(valid_parameter_ids):
        raise HTTPException(status_code=422, detail="A parameter does not belong to this LIS test")
    mapping = db.scalar(
        select(AnalyzerTestMapping).where(
            AnalyzerTestMapping.analyzer_id == analyzer.id,
            AnalyzerTestMapping.test_id == test.id,
        )
    )
    previous = None
    if mapping:
        previous = {
            "machine_test_code": mapping.machine_test_code,
            "parameter_count": len(mapping.parameters),
        }
        mapping.machine_test_code = payload.machine_test_code.strip().upper()
        mapping.status = Status.active
        mapping.updated_by = context.user_id
        mapping.parameters.clear()
        flush(db)
        response.status_code = status.HTTP_200_OK
        event_type = "analyzer.mapping_updated"
        action = "update"
    else:
        mapping = AnalyzerTestMapping(
            organization_id=context.organization_id,
            analyzer_id=analyzer.id,
            test_id=test.id,
            machine_test_code=payload.machine_test_code.strip().upper(),
            status=Status.active,
            created_by=context.user_id,
            updated_by=context.user_id,
        )
        db.add(mapping)
        event_type = "analyzer.mapping_created"
        action = "create"
    mapping.parameters.extend(
        AnalyzerParameterMapping(
            parameter_id=item.parameter_id,
            machine_parameter_code=item.machine_parameter_code.strip().upper(),
            unit=item.unit.strip() if item.unit else None,
        )
        for item in payload.parameters
    )
    flush(db)
    record_event(
        db,
        request,
        context,
        event_type=event_type,
        entity_type="analyzer_test_mapping",
        entity_id=mapping.id,
        branch_id=analyzer.branch_id,
        action=action,
        previous=previous,
        new={
            "lis_test_code": test.code,
            "machine_test_code": mapping.machine_test_code,
            "parameter_count": len(mapping.parameters),
        },
    )
    commit(db)
    return analyzer_mapping_read(db, mapping)


@router.delete("/analyzers/{analyzer_id}/mappings/{mapping_id}", status_code=204)
def delete_analyzer_mapping(
    analyzer_id: uuid.UUID,
    mapping_id: uuid.UUID,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.manage"))],
) -> Response:
    analyzer = get_tenant_record(db, Analyzer, analyzer_id, context)
    if not context.can_access_branch(analyzer.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    mapping = db.scalar(
        select(AnalyzerTestMapping).where(
            AnalyzerTestMapping.id == mapping_id,
            AnalyzerTestMapping.analyzer_id == analyzer.id,
            AnalyzerTestMapping.organization_id == context.organization_id,
        )
    )
    if mapping is None:
        raise HTTPException(status_code=404, detail="Mapping not found")
    test = db.get(TestCatalogItem, mapping.test_id)
    previous = {
        "lis_test_code": test.code if test else None,
        "machine_test_code": mapping.machine_test_code,
        "status": mapping.status.value if hasattr(mapping.status, "value") else mapping.status,
    }
    record_event(
        db,
        request,
        context,
        event_type="analyzer.mapping_deleted",
        entity_type="analyzer_test_mapping",
        entity_id=mapping.id,
        branch_id=analyzer.branch_id,
        action="delete",
        previous=previous,
    )
    db.delete(mapping)
    commit(db)
    return Response(status_code=204)


@router.patch(
    "/analyzers/{analyzer_id}/mappings/{mapping_id}",
    response_model=AnalyzerTestMappingRead,
)
def update_analyzer_mapping_status(
    analyzer_id: uuid.UUID,
    mapping_id: uuid.UUID,
    payload: AnalyzerMappingStatusUpdate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.manage"))],
) -> AnalyzerTestMappingRead:
    analyzer = get_tenant_record(db, Analyzer, analyzer_id, context)
    if not context.can_access_branch(analyzer.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    mapping = db.scalar(
        select(AnalyzerTestMapping).where(
            AnalyzerTestMapping.id == mapping_id,
            AnalyzerTestMapping.analyzer_id == analyzer.id,
            AnalyzerTestMapping.organization_id == context.organization_id,
        )
    )
    if mapping is None:
        raise HTTPException(status_code=404, detail="Mapping not found")
    previous = {
        "status": mapping.status.value if hasattr(mapping.status, "value") else mapping.status
    }
    mapping.status = payload.status
    mapping.updated_by = context.user_id
    record_event(
        db,
        request,
        context,
        event_type="analyzer.mapping_status_updated",
        entity_type="analyzer_test_mapping",
        entity_id=mapping.id,
        branch_id=analyzer.branch_id,
        action="update",
        previous=previous,
        new={"status": payload.status.value},
    )
    commit(db)
    return analyzer_mapping_read(db, mapping)


@router.get("/organizations", response_model=Page[OrganizationRead])
def list_organizations(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("organization.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
) -> Page[OrganizationRead]:
    statement = (
        select(Organization)
        .where(Organization.id == context.organization_id)
        .order_by(Organization.code, Organization.id)
    )
    return page(db, statement, OrganizationRead, limit, offset)


@router.post("/organizations", response_model=OrganizationRead, status_code=201)
def create_organization(
    payload: OrganizationCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("organization.manage"))],
) -> Organization:
    if db.scalar(select(Organization).where(Organization.id == context.organization_id)):
        raise HTTPException(
            status_code=403, detail="Tenant users cannot create another organization"
        )
    organization = Organization(
        **payload.model_dump(), created_by=context.user_id, updated_by=context.user_id
    )
    db.add(organization)
    flush(db)
    record_event(
        db,
        request,
        context,
        event_type="organization.created",
        entity_type="organization",
        entity_id=organization.id,
        action="create",
        new=payload.model_dump(),
    )
    commit(db)
    return organization


@router.get("/organizations/{organization_id}", response_model=OrganizationRead)
def get_organization(
    organization_id: uuid.UUID,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("organization.read"))],
) -> Organization:
    if organization_id != context.organization_id:
        raise HTTPException(status_code=404, detail="Organization not found")
    organization = db.get(Organization, organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Organization not found")
    return organization


@router.patch("/organizations/{organization_id}", response_model=OrganizationRead)
def update_organization(
    organization_id: uuid.UUID,
    payload: OrganizationUpdate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("organization.manage"))],
) -> Organization:
    organization = get_organization(organization_id, db, context)
    previous = snapshot(organization, ["name", "status"])
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(organization, key, value)
    organization.updated_by = context.user_id
    record_event(
        db,
        request,
        context,
        event_type="organization.updated",
        entity_type="organization",
        entity_id=organization.id,
        action="update",
        previous=previous,
        new=snapshot(organization, ["name", "status"]),
    )
    commit(db)
    return organization


@router.get("/branches", response_model=Page[BranchRead])
def list_branches(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
) -> Page[BranchRead]:
    statement = select(Branch).where(Branch.organization_id == context.organization_id)
    if not context.is_organization_scoped:
        statement = statement.where(Branch.id.in_(context.branch_ids))
    return page(db, statement.order_by(Branch.code, Branch.id), BranchRead, limit, offset)


@router.post("/branches", response_model=BranchRead, status_code=201)
def create_branch(
    payload: BranchCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.manage"))],
) -> Branch:
    if not context.is_organization_scoped:
        raise HTTPException(status_code=403, detail="Organization-scoped role required")
    branch = Branch(
        **payload.model_dump(),
        organization_id=context.organization_id,
        created_by=context.user_id,
        updated_by=context.user_id,
    )
    db.add(branch)
    flush(db)
    record_event(
        db,
        request,
        context,
        event_type="branch.created",
        entity_type="branch",
        entity_id=branch.id,
        branch_id=branch.id,
        action="create",
        new=payload.model_dump(),
    )
    commit(db)
    return branch


@router.get("/branches/{branch_id}", response_model=BranchRead)
def get_branch(
    branch_id: uuid.UUID,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.read"))],
) -> Branch:
    branch = get_tenant_record(db, Branch, branch_id, context)
    if not context.can_access_branch(branch.id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    return branch


@router.patch("/branches/{branch_id}", response_model=BranchRead)
def update_branch(
    branch_id: uuid.UUID,
    payload: BranchUpdate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.manage"))],
) -> Branch:
    branch = get_branch(branch_id, db, context)
    previous = snapshot(branch, ["name", "address", "time_zone", "status"])
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(branch, key, value)
    branch.updated_by = context.user_id
    record_event(
        db,
        request,
        context,
        event_type="branch.updated",
        entity_type="branch",
        entity_id=branch.id,
        branch_id=branch.id,
        action="update",
        previous=previous,
        new=snapshot(branch, ["name", "address", "time_zone", "status"]),
    )
    commit(db)
    return branch


@router.get("/departments", response_model=Page[DepartmentRead])
def list_departments(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
) -> Page[DepartmentRead]:
    statement = select(Department).where(Department.organization_id == context.organization_id)
    if not context.is_organization_scoped:
        statement = statement.where(Department.branch_id.in_(context.branch_ids))
    return page(
        db, statement.order_by(Department.code, Department.id), DepartmentRead, limit, offset
    )


@router.post("/departments", response_model=DepartmentRead, status_code=201)
def create_department(
    payload: DepartmentCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.manage"))],
) -> Department:
    if payload.branch_id:
        get_branch(payload.branch_id, db, context)
    elif not context.is_organization_scoped:
        raise HTTPException(status_code=403, detail="Organization-scoped role required")
    department = Department(**payload.model_dump(), organization_id=context.organization_id)
    db.add(department)
    flush(db)
    record_event(
        db,
        request,
        context,
        event_type="department.created",
        entity_type="department",
        entity_id=department.id,
        branch_id=department.branch_id,
        action="create",
        new=payload.model_dump(mode="json"),
    )
    commit(db)
    return department


@router.get("/departments/{department_id}", response_model=DepartmentRead)
def get_department(
    department_id: uuid.UUID,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.read"))],
) -> Department:
    department = get_tenant_record(db, Department, department_id, context)
    if not context.can_access_branch(department.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    return department


@router.patch("/departments/{department_id}", response_model=DepartmentRead)
def update_department(
    department_id: uuid.UUID,
    payload: DepartmentUpdate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.manage"))],
) -> Department:
    department = get_department(department_id, db, context)
    previous = snapshot(department, ["name", "status"])
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(department, key, value)
    record_event(
        db,
        request,
        context,
        event_type="department.updated",
        entity_type="department",
        entity_id=department.id,
        branch_id=department.branch_id,
        action="update",
        previous=previous,
        new=snapshot(department, ["name", "status"]),
    )
    commit(db)
    return department


@router.get("/users", response_model=Page[UserRead])
def list_users(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("user.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
) -> Page[UserRead]:
    statement = (
        select(User)
        .where(User.organization_id == context.organization_id)
        .order_by(User.email, User.id)
    )
    return page(db, statement, UserRead, limit, offset)


@router.post("/users", response_model=UserRead, status_code=201)
def create_user(
    payload: UserCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("user.manage"))],
) -> User:
    user = User(**payload.model_dump(), organization_id=context.organization_id)
    db.add(user)
    flush(db)
    record_event(
        db,
        request,
        context,
        event_type="user.created",
        entity_type="user",
        entity_id=user.id,
        action="create",
        new=payload.model_dump(),
    )
    commit(db)
    return user


@router.get("/users/{user_id}", response_model=UserRead)
def get_user(
    user_id: uuid.UUID,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("user.read"))],
) -> User:
    return get_tenant_record(db, User, user_id, context)


@router.patch("/users/{user_id}", response_model=UserRead)
def update_user(
    user_id: uuid.UUID,
    payload: UserUpdate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("user.manage"))],
) -> User:
    user = get_user(user_id, db, context)
    previous = snapshot(user, ["display_name", "status"])
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(user, key, value)
    event_type = "user.status_changed" if payload.status is not None else "user.updated"
    record_event(
        db,
        request,
        context,
        event_type=event_type,
        entity_type="user",
        entity_id=user.id,
        action="update",
        previous=previous,
        new=snapshot(user, ["display_name", "status"]),
    )
    commit(db)
    return user


@router.get("/roles", response_model=Page[RoleRead])
def list_roles(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("role.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
) -> Page[RoleRead]:
    statement = (
        select(Role)
        .where((Role.organization_id == context.organization_id) | (Role.organization_id.is_(None)))
        .order_by(Role.name, Role.id)
    )
    return page(db, statement, RoleRead, limit, offset)


@router.post("/roles", response_model=RoleRead, status_code=201)
def create_role(
    payload: RoleCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("role.manage"))],
) -> Role:
    permissions = list(
        db.scalars(select(Permission).where(Permission.code.in_(payload.permission_codes))).all()
    )
    if len(permissions) != len(set(payload.permission_codes)):
        raise HTTPException(status_code=422, detail="One or more permissions are unknown")
    role = Role(
        organization_id=context.organization_id,
        name=payload.name,
        description=payload.description,
        permissions=permissions,
    )
    db.add(role)
    flush(db)
    record_event(
        db,
        request,
        context,
        event_type="role.created",
        entity_type="role",
        entity_id=role.id,
        action="create",
        new=payload.model_dump(),
    )
    commit(db)
    return role


@router.get("/permissions", response_model=list[PermissionRead])
def list_permissions(
    db: Db, _: Annotated[AuthContext, Depends(require_permission("role.read"))]
) -> list[Permission]:
    return list(db.scalars(select(Permission).order_by(Permission.code)).all())


@router.post("/user-role-assignments", response_model=AssignmentRead, status_code=201)
def assign_role(
    payload: AssignmentCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("role.manage"))],
) -> UserRoleAssignment:
    get_user(payload.user_id, db, context)
    role = db.scalar(
        select(Role).where(
            Role.id == payload.role_id,
            (Role.organization_id == context.organization_id) | (Role.organization_id.is_(None)),
        )
    )
    if role is None:
        raise HTTPException(status_code=404, detail="Role not found")
    if payload.branch_id:
        get_branch(payload.branch_id, db, context)
    assignment_data = payload.model_dump(exclude={"effective_at"})
    assignment = UserRoleAssignment(
        **assignment_data,
        organization_id=context.organization_id,
        assigned_by=context.user_id,
        effective_at=payload.effective_at or datetime.now(UTC),
    )
    db.add(assignment)
    flush(db)
    record_event(
        db,
        request,
        context,
        event_type="role.assigned",
        entity_type="user_role_assignment",
        entity_id=assignment.id,
        branch_id=assignment.branch_id,
        action="assign",
        new=payload.model_dump(mode="json"),
    )
    commit(db)
    return assignment


@router.delete("/user-role-assignments/{assignment_id}", status_code=204)
def revoke_role(
    assignment_id: uuid.UUID,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("role.manage"))],
) -> Response:
    assignment = get_tenant_record(db, UserRoleAssignment, assignment_id, context)
    if not context.can_access_branch(assignment.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    assignment.active = False
    record_event(
        db,
        request,
        context,
        event_type="role.revoked",
        entity_type="user_role_assignment",
        entity_id=assignment.id,
        branch_id=assignment.branch_id,
        action="deactivate",
        previous={"active": True},
        new={"active": False},
    )
    commit(db)
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/audit-events", response_model=Page[AuditEventRead])
def list_audit_events(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("audit.read"))],
    event_type: str | None = None,
    entity_type: str | None = None,
    branch_id: uuid.UUID | None = None,
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
) -> Page[AuditEventRead]:
    statement = select(AuditEvent).where(AuditEvent.organization_id == context.organization_id)
    if not context.is_organization_scoped:
        statement = statement.where(AuditEvent.branch_id.in_(context.branch_ids))
    if branch_id:
        if not context.can_access_branch(branch_id):
            raise HTTPException(status_code=403, detail="Branch access denied")
        statement = statement.where(AuditEvent.branch_id == branch_id)
    if event_type:
        statement = statement.where(AuditEvent.event_type == event_type)
    if entity_type:
        statement = statement.where(AuditEvent.entity_type == entity_type)
    statement = statement.order_by(AuditEvent.occurred_at.desc(), AuditEvent.id.desc())
    return page(db, statement, AuditEventRead, limit, offset)


@router.get("/audit-events/{event_id}", response_model=AuditEventRead)
def get_audit_event(
    event_id: uuid.UUID,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("audit.read"))],
) -> AuditEvent:
    event_record = get_tenant_record(db, AuditEvent, event_id, context)
    if not context.can_access_branch(event_record.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    return event_record


@router.get("/test-catalog", response_model=list[TestCatalogRead])
def list_test_catalog(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.read"))],
) -> list[TestCatalogItem]:
    return list(
        db.scalars(
            select(TestCatalogItem)
            .where(
                TestCatalogItem.organization_id == context.organization_id,
                TestCatalogItem.status == "active",
            )
            .order_by(TestCatalogItem.name, TestCatalogItem.id)
        ).all()
    )


@router.get("/test-master", response_model=Page[TestMasterRead])
def list_test_master(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("test_master.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
    search: Annotated[str | None, Query(max_length=200)] = None,
    review_only: bool = False,
) -> Page[TestMasterRead]:
    statement = select(TestCatalogItem).where(
        TestCatalogItem.organization_id == context.organization_id
    )
    if search:
        term = f"%{search.strip().lower()}%"
        statement = statement.where(
            or_(
                func.lower(TestCatalogItem.code).like(term),
                func.lower(TestCatalogItem.name).like(term),
                func.lower(TestCatalogItem.sub_department).like(term),
                func.lower(TestCatalogItem.specimen_type).like(term),
            )
        )
    if review_only:
        statement = statement.where(TestCatalogItem.validation_status == "needs_review")
    return page(
        db,
        statement.order_by(TestCatalogItem.name, TestCatalogItem.id),
        TestMasterRead,
        limit,
        offset,
    )


@router.post("/test-master", response_model=TestMasterRead, status_code=201)
def create_test_master_item(
    payload: TestMasterCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("test_master.manage"))],
) -> TestCatalogItem:
    item = TestCatalogItem(
        **payload.model_dump(),
        organization_id=context.organization_id,
        is_panel=False,
        validation_status=(
            "needs_review" if payload.specimen_type.strip().lower() == "specimen" else "validated"
        ),
    )
    db.add(item)
    flush(db)
    record_event(
        db,
        request,
        context,
        event_type="test_master.created",
        entity_type="test_master",
        entity_id=item.id,
        action="create",
        new=payload.model_dump(mode="json"),
    )
    commit(db)
    return item


@router.post(
    "/test-master/{test_id}/parameters",
    response_model=TestParameterRead,
    status_code=201,
)
def create_test_parameter(
    test_id: uuid.UUID,
    payload: TestParameterCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("test_master.manage"))],
) -> TestCatalogParameter:
    test = get_tenant_record(db, TestCatalogItem, test_id, context)
    parameter = TestCatalogParameter(
        test_id=test.id,
        name=payload.name.strip(),
        external_code=payload.external_code.strip().upper(),
        display_order=payload.display_order,
        unit=payload.unit.strip() if payload.unit else None,
        reference_low=payload.reference_low.strip() if payload.reference_low else None,
        reference_high=payload.reference_high.strip() if payload.reference_high else None,
        reference_text=payload.reference_text.strip() if payload.reference_text else None,
    )
    test.parameters.append(parameter)
    test.is_panel = True
    has_reference = bool(
        parameter.reference_low or parameter.reference_high or parameter.reference_text
    )
    if not has_reference:
        test.validation_status = "needs_review"
    flush(db)
    record_event(
        db,
        request,
        context,
        event_type="test_master.parameter_created",
        entity_type="test_catalog_parameter",
        entity_id=parameter.id,
        action="create",
        new=payload.model_dump(mode="json"),
    )
    commit(db)
    return parameter


@router.patch(
    "/test-master/{test_id}/parameters/{parameter_id}",
    response_model=TestParameterRead,
)
def update_test_parameter(
    test_id: uuid.UUID,
    parameter_id: uuid.UUID,
    payload: TestParameterUpdate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("test_master.manage"))],
) -> TestCatalogParameter:
    test = get_tenant_record(db, TestCatalogItem, test_id, context)
    parameter = db.scalar(
        select(TestCatalogParameter).where(
            TestCatalogParameter.id == parameter_id,
            TestCatalogParameter.test_id == test.id,
        )
    )
    if parameter is None:
        raise HTTPException(status_code=404, detail="Parameter not found")
    previous = {
        "name": parameter.name,
        "external_code": parameter.external_code,
        "unit": parameter.unit,
        "reference_low": parameter.reference_low,
        "reference_high": parameter.reference_high,
        "reference_text": parameter.reference_text,
    }
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        if isinstance(value, str):
            value = value.strip()
            if key == "external_code":
                value = value.upper()
            if value == "":
                value = None
        setattr(parameter, key, value)
    record_event(
        db,
        request,
        context,
        event_type="test_master.parameter_updated",
        entity_type="test_catalog_parameter",
        entity_id=parameter.id,
        action="update",
        previous=previous,
        new=data,
    )
    commit(db)
    return parameter


@router.post("/test-master/import", response_model=TestMasterImportRead)
async def import_test_master(
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("test_master.manage"))],
    file: Annotated[UploadFile, File(description="LIS/HIS Test Master .xlsx workbook")],
) -> TestMasterImportRead:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Upload an .xlsx test master workbook")
    from io import BytesIO

    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(await file.read()), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip().upper() for value in next(rows)]
    except Exception as error:
        raise HTTPException(status_code=422, detail="The workbook could not be read") from error
    required = {
        "SERVICE TYPE",
        "DEPARTMENT",
        "SUB DEPARTMENT",
        "SERVICE CODE",
        "SERVICE NAME",
        "SPECIMEN",
        "PARAMETER CODE",
        "PARAMETER DESCRIPTION",
    }
    if not required.issubset(headers):
        missing = ", ".join(sorted(required.difference(headers)))
        raise HTTPException(status_code=422, detail=f"Missing columns: {missing}")
    positions = {name: headers.index(name) for name in required}
    grouped: dict[str, dict[str, Any]] = {}
    rejected = 0
    errors: list[str] = []
    row_count = 0

    def cell_value(row: tuple[Any, ...], key: str) -> str:
        return str(row[positions[key]] or "").strip()

    for row_number, row in enumerate(rows, start=2):
        row_count += 1
        code = cell_value(row, "SERVICE CODE")
        name = cell_value(row, "SERVICE NAME")
        if not code or not name:
            rejected += 1
            if len(errors) < 20:
                errors.append(f"Row {row_number}: service code and service name are required")
            continue
        entry = grouped.setdefault(
            code,
            {
                "name": name,
                "service_type": cell_value(row, "SERVICE TYPE") or "Pathology",
                "department": cell_value(row, "DEPARTMENT") or "Laboratory",
                "sub_department": cell_value(row, "SUB DEPARTMENT"),
                "specimen_type": cell_value(row, "SPECIMEN") or "specimen",
                "parameters": [],
            },
        )
        parameter_name = cell_value(row, "PARAMETER CODE")
        external_code = cell_value(row, "PARAMETER DESCRIPTION")
        if parameter_name or external_code:
            if parameter_name and external_code:
                entry["parameters"].append((parameter_name, external_code))
            else:
                rejected += 1
                if len(errors) < 20:
                    errors.append(
                        f"Row {row_number}: parameter name and identifier must both be present"
                    )
    created = updated = parameter_count = review_required = 0
    for code, entry in grouped.items():
        item = db.scalar(
            select(TestCatalogItem).where(
                TestCatalogItem.organization_id == context.organization_id,
                TestCatalogItem.code == code,
            )
        )
        needs_review = entry["specimen_type"].lower() == "specimen"
        if item is None:
            item = TestCatalogItem(
                organization_id=context.organization_id,
                code=code,
                container_type="Unspecified",
                price=Decimal("0"),
            )
            db.add(item)
            created += 1
        else:
            updated += 1
            item.parameters.clear()
        for key in ("name", "service_type", "department", "sub_department", "specimen_type"):
            setattr(item, key, entry[key])
        item.is_panel = bool(entry["parameters"])
        item.validation_status = "needs_review" if needs_review else "validated"
        if needs_review:
            review_required += 1
        seen: set[str] = set()
        for order, (parameter_name, external_code) in enumerate(entry["parameters"]):
            if external_code in seen:
                continue
            seen.add(external_code)
            item.parameters.append(
                TestCatalogParameter(
                    name=parameter_name, external_code=external_code, display_order=order
                )
            )
            parameter_count += 1
    summary = TestMasterImportRead(
        rows_received=row_count,
        tests_created=created,
        tests_updated=updated,
        parameters_imported=parameter_count,
        rows_rejected=rejected,
        review_required=review_required,
        errors=errors,
    )
    record_event(
        db,
        request,
        context,
        event_type="test_master.imported",
        entity_type="test_master",
        entity_id=None,
        action="import",
        new={**summary.model_dump(), "filename": file.filename},
    )
    commit(db)
    return summary


@router.get("/patients/lookup", response_model=PatientLookupRead | None)
def lookup_patient(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.read"))],
    query: Annotated[str, Query(min_length=3, max_length=320)],
) -> PatientLookupRead | None:
    """Find a returning patient by UUID, patient number, phone, email, or name."""
    value = query.strip()
    conditions = [
        Patient.patient_number == value,
        func.lower(Patient.email) == value.lower(),
        Patient.phone == value,
    ]
    try:
        conditions.append(Patient.id == uuid.UUID(value))
    except ValueError:
        pass
    patient = db.scalar(
        select(Patient).where(
            Patient.organization_id == context.organization_id,
            or_(*conditions),
        )
    )
    if patient is None:
        # Prefer an exact name match; otherwise allow a unique partial name match.
        name_filter = func.lower(Patient.full_name) == value.lower()
        name_matches = list(
            db.scalars(
                select(Patient)
                .where(
                    Patient.organization_id == context.organization_id,
                    name_filter,
                )
                .order_by(Patient.updated_at.desc(), Patient.id.desc())
                .limit(2)
            ).all()
        )
        if not name_matches:
            name_filter = func.lower(Patient.full_name).like(f"%{value.lower()}%")
            name_matches = list(
                db.scalars(
                    select(Patient)
                    .where(
                        Patient.organization_id == context.organization_id,
                        name_filter,
                    )
                    .order_by(Patient.updated_at.desc(), Patient.id.desc())
                    .limit(2)
                ).all()
            )
        if len(name_matches) > 1:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Multiple patients match this name. Search using the full name, "
                    "patient number, phone, email, or UUID."
                ),
            )
        patient = name_matches[0] if name_matches else None
    if patient is None:
        return None
    visit_count, last_visit_at = db.execute(
        select(func.count(LabOrder.id), func.max(LabOrder.created_at)).where(
            LabOrder.organization_id == context.organization_id,
            LabOrder.patient_id == patient.id,
        )
    ).one()
    return PatientLookupRead(
        **{
            key: getattr(patient, key)
            for key in (
                "id",
                "patient_number",
                "full_name",
                "phone",
                "email",
                "date_of_birth",
                "age_years",
                "sex",
                "address",
                "blood_group",
                "country",
                "race",
                "nationality",
                "additional_patient_data",
            )
        },
        visit_count=visit_count,
        last_visit_at=last_visit_at,
    )


@router.get("/patients", response_model=Page[PatientLookupRead])
def list_patients(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
    search: Annotated[str | None, Query(max_length=200)] = None,
) -> Page[PatientLookupRead]:
    visit_summary = (
        select(
            LabOrder.patient_id.label("patient_id"),
            func.count(LabOrder.id).label("visit_count"),
            func.max(LabOrder.created_at).label("last_visit_at"),
        )
        .where(LabOrder.organization_id == context.organization_id)
        .group_by(LabOrder.patient_id)
        .subquery()
    )
    filters = [Patient.organization_id == context.organization_id]
    if search:
        term = f"%{search.strip().lower()}%"
        filters.append(
            or_(
                func.lower(Patient.patient_number).like(term),
                func.lower(Patient.full_name).like(term),
                func.lower(Patient.phone).like(term),
                func.lower(Patient.email).like(term),
            )
        )
    rows = db.execute(
        select(
            Patient,
            func.coalesce(visit_summary.c.visit_count, 0),
            visit_summary.c.last_visit_at,
            func.count(Patient.id).over().label("total_count"),
        )
        .outerjoin(visit_summary, visit_summary.c.patient_id == Patient.id)
        .where(*filters)
        .order_by(Patient.updated_at.desc(), Patient.id.desc())
        .limit(limit)
        .offset(offset)
    ).all()
    total = rows[0].total_count if rows else 0
    if not rows and offset:
        total = db.scalar(select(func.count(Patient.id)).where(*filters)) or 0
    items: list[PatientLookupRead] = []
    for patient, visit_count, last_visit_at, _ in rows:
        items.append(
            PatientLookupRead.model_validate(
                {
                    **snapshot(
                        patient,
                        [
                            "id",
                            "patient_number",
                            "full_name",
                            "phone",
                            "email",
                            "date_of_birth",
                            "age_years",
                            "sex",
                            "address",
                            "blood_group",
                            "country",
                            "race",
                            "nationality",
                            "additional_patient_data",
                        ],
                    ),
                    "visit_count": visit_count,
                    "last_visit_at": last_visit_at,
                }
            )
        )
    return Page[PatientLookupRead](items=items, total=total, limit=limit, offset=offset)


@router.post("/intake-workflows", response_model=IntakeRead, status_code=201)
def create_intake_workflow(
    payload: IntakeCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.manage"))],
) -> IntakeRead:
    branch_statement = select(Branch).where(Branch.organization_id == context.organization_id)
    if not context.is_organization_scoped:
        branch_statement = branch_statement.where(Branch.id.in_(context.branch_ids))
    branch = db.scalar(branch_statement.order_by(Branch.code))
    if branch is None:
        raise HTTPException(status_code=422, detail="No accessible branch is configured")

    catalog_items = list(
        db.scalars(
            select(TestCatalogItem).where(
                TestCatalogItem.organization_id == context.organization_id,
                TestCatalogItem.id.in_(payload.test_ids),
                TestCatalogItem.status == "active",
            )
        ).all()
    )
    if len(catalog_items) != len(set(payload.test_ids)):
        raise HTTPException(status_code=422, detail="One or more tests are unavailable")

    stamp = datetime.now(UTC).strftime("%Y%m%d%H%M%S")
    suffix = uuid.uuid4().hex[:6].upper()
    normalized_email = str(payload.email).strip().lower()
    patient = None
    if payload.patient_id:
        patient = db.scalar(
            select(Patient).where(
                Patient.id == payload.patient_id,
                Patient.organization_id == context.organization_id,
            )
        )
        if patient is None:
            raise HTTPException(status_code=404, detail="Patient not found")
    else:
        matches = list(
            db.scalars(
                select(Patient).where(
                    Patient.organization_id == context.organization_id,
                    or_(
                        Patient.phone == payload.phone.strip(),
                        func.lower(Patient.email) == normalized_email,
                    ),
                )
            ).all()
        )
        if len({item.id for item in matches}) > 1:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Phone and email match different patients; search and select "
                    "the correct patient"
                ),
            )
        patient = matches[0] if matches else None

    optional_patient_data = {
        key: value.strip()
        for key, value in payload.additional_patient_data.items()
        if value.strip()
    }
    patient_values = {
        "full_name": payload.full_name.strip() if payload.full_name else "Unknown patient",
        "date_of_birth": payload.date_of_birth,
        "age_years": payload.age_years,
        "sex": payload.sex,
        "phone": payload.phone.strip(),
        "email": normalized_email,
        "address": payload.address,
        "blood_group": payload.blood_group,
        "country": payload.country,
        "race": payload.race,
        "nationality": payload.nationality,
        "additional_patient_data": optional_patient_data,
    }
    patient_created = patient is None
    if patient is None:
        patient = Patient(
            organization_id=context.organization_id,
            patient_number=f"PT-{stamp}-{suffix}",
            **patient_values,
        )
        db.add(patient)
        flush(db)
    else:
        for key, value in patient_values.items():
            setattr(patient, key, value)
    order = LabOrder(
        organization_id=context.organization_id,
        branch_id=branch.id,
        patient_id=patient.id,
        order_number=f"ORD-{stamp}-{suffix}",
        visit_type=payload.visit_type,
        department=payload.department,
        ward=payload.ward,
        doctor_name=payload.doctor_name,
        diagnosis=payload.diagnosis,
        prescription_filename=payload.prescription_filename,
        notes=payload.notes,
    )
    db.add(order)
    flush(db)
    db.add(
        PatientHistory(
            organization_id=context.organization_id,
            patient_id=patient.id,
            order_id=order.id,
            recorded_by=context.user_id,
            demographics={
                **{
                    key: (value.isoformat() if isinstance(value, date) else value)
                    for key, value in patient_values.items()
                },
                "visit_type": payload.visit_type,
                "department": payload.department,
                "ward": payload.ward,
                "doctor_name": payload.doctor_name,
                "diagnosis": payload.diagnosis,
            },
        )
    )

    subtotal = sum((Decimal(str(item.price)) for item in catalog_items), start=Decimal("0"))
    if payload.discount > subtotal:
        raise HTTPException(status_code=422, detail="Discount cannot exceed subtotal")
    for item in catalog_items:
        db.add(OrderTest(order_id=order.id, test_id=item.id, price=item.price))
    invoice = Invoice(
        organization_id=context.organization_id,
        order_id=order.id,
        invoice_number=f"INV-{stamp}-{suffix}",
        subtotal=subtotal,
        discount=payload.discount,
        total=subtotal - payload.discount,
    )
    db.add(invoice)

    record_event(
        db,
        request,
        context,
        event_type="intake.registered",
        entity_type="lab_order",
        entity_id=order.id,
        branch_id=branch.id,
        action="create",
        new={
            "patient_number": patient.patient_number,
            "order_number": order.order_number,
            "test_count": len(catalog_items),
            "invoice_number": invoice.invoice_number,
            "payment_status": "pending",
            "patient_created": patient_created,
        },
    )
    commit(db)
    return IntakeRead(
        patient_id=patient.id,
        patient_number=patient.patient_number,
        order_id=order.id,
        order_number=order.order_number,
        invoice_number=invoice.invoice_number,
        subtotal=invoice.subtotal,
        discount=invoice.discount,
        total=invoice.total,
        payment_status=invoice.payment_status,
        specimens=[],
    )


def payment_summary(
    db: Session, order_id: uuid.UUID, context: AuthContext
) -> tuple[LabOrder, Patient, Invoice]:
    order = get_tenant_record(db, LabOrder, order_id, context)
    if not context.can_access_branch(order.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    patient = db.get(Patient, order.patient_id)
    invoice = db.scalar(
        select(Invoice).where(
            Invoice.organization_id == context.organization_id,
            Invoice.order_id == order.id,
        )
    )
    if patient is None or invoice is None:
        raise HTTPException(status_code=404, detail="Payment record not found")
    return order, patient, invoice


@router.get("/orders/{order_id}/payment", response_model=PaymentSummary)
def get_payment_summary(
    order_id: uuid.UUID,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.read"))],
) -> PaymentSummary:
    order, patient, invoice = payment_summary(db, order_id, context)
    return PaymentSummary(
        order_id=order.id,
        order_number=order.order_number,
        patient_number=patient.patient_number,
        patient_name=patient.full_name,
        invoice_number=invoice.invoice_number,
        total=invoice.total,
        payment_status=invoice.payment_status,
    )


@router.post("/orders/{order_id}/payment", response_model=PaymentRead)
def record_payment(
    order_id: uuid.UUID,
    payload: PaymentCreate,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.manage"))],
) -> PaymentRead:
    order, patient, invoice = payment_summary(db, order_id, context)
    transaction_id = payload.transaction_id.strip() if payload.transaction_id else None
    if payload.payment_method in {"UPI", "CARD"} and not transaction_id:
        raise HTTPException(status_code=422, detail="Transaction ID is required for UPI or card")
    if payload.payment_method == "CASH":
        transaction_id = None

    if invoice.payment_status != "paid":
        invoice.payment_status = "paid"
        invoice.payment_method = payload.payment_method
        invoice.transaction_id = transaction_id
        invoice.paid_at = datetime.now(UTC)
        order.status = "awaiting_collection"

        catalog_items = list(
            db.scalars(
                select(TestCatalogItem)
                .join(OrderTest, OrderTest.test_id == TestCatalogItem.id)
                .where(OrderTest.order_id == order.id)
            ).all()
        )
        specimen_groups: dict[tuple[str, str], set[str]] = {}
        for item in catalog_items:
            specimen_groups.setdefault((item.specimen_type, item.container_type), set()).add(
                item.sub_department or "Unassigned"
            )
        suffix = order.order_number.rsplit("-", 1)[-1]
        stamp = datetime.now(UTC).strftime("%m%d%H%M%S")
        for index, ((specimen_type, container_type), departments) in enumerate(
            sorted(specimen_groups.items()), start=1
        ):
            db.add(
                Specimen(
                    organization_id=context.organization_id,
                    branch_id=order.branch_id,
                    order_id=order.id,
                    barcode=f"LQ{stamp}{suffix}{index:02d}",
                    specimen_type=specimen_type,
                    container_type=container_type,
                    laboratory_department=", ".join(sorted(departments)),
                )
            )
        record_event(
            db,
            request,
            context,
            event_type="payment.recorded",
            entity_type="invoice",
            entity_id=invoice.id,
            branch_id=order.branch_id,
            action="pay",
            previous={"payment_status": "pending"},
            new={
                "payment_status": "paid",
                "payment_method": payload.payment_method,
                "transaction_id": transaction_id,
            },
        )
        commit(db)

    specimens = list(
        db.scalars(
            select(Specimen)
            .where(
                Specimen.organization_id == context.organization_id,
                Specimen.order_id == order.id,
            )
            .order_by(Specimen.barcode)
        ).all()
    )
    assert invoice.paid_at is not None
    assert invoice.payment_method is not None
    return PaymentRead(
        order_id=order.id,
        order_number=order.order_number,
        patient_number=patient.patient_number,
        patient_name=patient.full_name,
        invoice_number=invoice.invoice_number,
        total=invoice.total,
        payment_status=invoice.payment_status,
        payment_method=invoice.payment_method,
        transaction_id=invoice.transaction_id,
        paid_at=invoice.paid_at,
        specimens=[SpecimenRead.model_validate(item) for item in specimens],
    )


def specimen_workflow_read(db: Session, specimen: Specimen) -> SpecimenWorkflowRead:
    order = db.get(LabOrder, specimen.order_id)
    patient = db.get(Patient, order.patient_id) if order else None
    if order is None or patient is None:
        raise HTTPException(status_code=404, detail="Specimen order was not found")
    return SpecimenWorkflowRead(
        **SpecimenRead.model_validate(specimen).model_dump(),
        id=specimen.id,
        order_id=order.id,
        order_number=order.order_number,
        patient_number=patient.patient_number,
        patient_name=patient.full_name,
        laboratory_department=specimen.laboratory_department,
        accession_number=specimen.accession_number,
        collection_location=specimen.collection_location,
        container_count=specimen.container_count,
        collection_notes=specimen.collection_notes,
        collected_at=specimen.collected_at,
        received_at=specimen.received_at,
        reviewed_at=specimen.reviewed_at,
        rejection_reason=specimen.rejection_reason,
        rejection_notes=specimen.rejection_notes,
    )


def get_specimen_by_barcode(db: Session, barcode: str, context: AuthContext) -> Specimen:
    specimen = db.scalar(
        select(Specimen).where(
            Specimen.organization_id == context.organization_id,
            Specimen.barcode == barcode.strip(),
        )
    )
    if specimen is None:
        raise HTTPException(status_code=404, detail="Specimen barcode not found")
    if not context.can_access_branch(specimen.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    return specimen


@router.get("/specimens", response_model=Page[SpecimenWorkflowRead])
def list_specimens(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 50,
    offset: Annotated[int, Query(ge=0)] = 0,
    specimen_status: Annotated[str | None, Query(alias="status")] = None,
    department: Annotated[str | None, Query(max_length=120)] = None,
    search: Annotated[str | None, Query(max_length=200)] = None,
) -> Page[SpecimenWorkflowRead]:
    statement = (
        select(Specimen)
        .join(LabOrder, LabOrder.id == Specimen.order_id)
        .join(Patient, Patient.id == LabOrder.patient_id)
        .where(Specimen.organization_id == context.organization_id)
    )
    if not context.is_organization_scoped:
        statement = statement.where(Specimen.branch_id.in_(context.branch_ids))
    if specimen_status:
        statement = statement.where(Specimen.status == specimen_status)
    if department:
        statement = statement.where(Specimen.laboratory_department == department)
    if search:
        term = f"%{search.strip().lower()}%"
        statement = statement.where(
            or_(
                func.lower(Specimen.barcode).like(term),
                func.lower(Specimen.accession_number).like(term),
                func.lower(LabOrder.order_number).like(term),
                func.lower(Patient.patient_number).like(term),
                func.lower(Patient.full_name).like(term),
            )
        )
    total = db.scalar(select(func.count()).select_from(statement.subquery())) or 0
    records = list(
        db.scalars(
            statement.order_by(Specimen.updated_at.desc(), Specimen.id.desc())
            .limit(limit)
            .offset(offset)
        ).all()
    )
    return Page[SpecimenWorkflowRead](
        items=[specimen_workflow_read(db, item) for item in records],
        total=total,
        limit=limit,
        offset=offset,
    )


@router.get("/specimens/{barcode}", response_model=SpecimenWorkflowRead)
def get_specimen(
    barcode: str,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.read"))],
) -> SpecimenWorkflowRead:
    return specimen_workflow_read(db, get_specimen_by_barcode(db, barcode, context))


@router.post("/specimens/{barcode}/collect", response_model=SpecimenWorkflowRead)
def collect_specimen(
    barcode: str,
    payload: SpecimenCollect,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.manage"))],
) -> SpecimenWorkflowRead:
    specimen = get_specimen_by_barcode(db, barcode, context)
    if specimen.status != "awaiting_collection":
        raise HTTPException(status_code=409, detail="Only awaiting specimens can be collected")
    specimen.status = "collected"
    specimen.collection_location = payload.collection_location.strip()
    specimen.container_count = payload.container_count
    specimen.collection_notes = payload.collection_notes
    specimen.collected_by = context.user_id
    specimen.collected_at = datetime.now(UTC)
    order = db.get(LabOrder, specimen.order_id)
    if order:
        order.status = "collection_in_progress"
    record_event(
        db,
        request,
        context,
        event_type="specimen.collected",
        entity_type="specimen",
        entity_id=specimen.id,
        branch_id=specimen.branch_id,
        action="collect",
        previous={"status": "awaiting_collection"},
        new={"status": "collected", **payload.model_dump()},
    )
    commit(db)
    return specimen_workflow_read(db, specimen)


@router.post("/specimens/{barcode}/receive", response_model=SpecimenWorkflowRead)
def receive_specimen(
    barcode: str,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.manage"))],
) -> SpecimenWorkflowRead:
    specimen = get_specimen_by_barcode(db, barcode, context)
    if specimen.status != "collected":
        raise HTTPException(status_code=409, detail="Only collected specimens can be received")
    specimen.status = "received"
    specimen.received_by = context.user_id
    specimen.received_at = datetime.now(UTC)
    specimen.accession_number = f"ACC-{datetime.now(UTC):%Y%m%d}-{uuid.uuid4().hex[:8].upper()}"
    order = db.get(LabOrder, specimen.order_id)
    if order:
        order.status = "laboratory_received"
    record_event(
        db,
        request,
        context,
        event_type="specimen.received",
        entity_type="specimen",
        entity_id=specimen.id,
        branch_id=specimen.branch_id,
        action="receive",
        previous={"status": "collected"},
        new={"status": "received", "accession_number": specimen.accession_number},
    )
    commit(db)
    return specimen_workflow_read(db, specimen)


@router.post("/specimens/{barcode}/decision", response_model=SpecimenWorkflowRead)
def review_specimen(
    barcode: str,
    payload: SpecimenDecision,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("branch.manage"))],
) -> SpecimenWorkflowRead:
    specimen = get_specimen_by_barcode(db, barcode, context)
    if specimen.status != "received":
        raise HTTPException(status_code=409, detail="Only received specimens can be reviewed")
    if payload.decision == "reject" and not payload.rejection_reason:
        raise HTTPException(status_code=422, detail="A rejection reason is required")
    specimen.status = "accepted" if payload.decision == "accept" else "rejected"
    specimen.reviewed_by = context.user_id
    specimen.reviewed_at = datetime.now(UTC)
    specimen.rejection_reason = payload.rejection_reason if payload.decision == "reject" else None
    specimen.rejection_notes = payload.notes
    order = db.get(LabOrder, specimen.order_id)
    if order:
        statuses = set(
            db.scalars(select(Specimen.status).where(Specimen.order_id == order.id)).all()
        )
        if "rejected" in statuses:
            order.status = "recollection_required"
        elif statuses == {"accepted"}:
            order.status = "ready_for_processing"
    worklist_created = 0
    if payload.decision == "accept":
        worklist_created = enqueue_analyzer_worklist_for_specimen(db, request, context, specimen)
    record_event(
        db,
        request,
        context,
        event_type=f"specimen.{specimen.status}",
        entity_type="specimen",
        entity_id=specimen.id,
        branch_id=specimen.branch_id,
        action=payload.decision,
        previous={"status": "received"},
        new={
            "status": specimen.status,
            "worklist_items_created": worklist_created,
            **payload.model_dump(),
        },
    )
    commit(db)
    return specimen_workflow_read(db, specimen)


def enqueue_analyzer_worklist_for_specimen(
    db: Session,
    request: Request,
    context: AuthContext,
    specimen: Specimen,
) -> int:
    order_tests = list(
        db.scalars(select(OrderTest).where(OrderTest.order_id == specimen.order_id)).all()
    )
    if not order_tests:
        return 0
    test_ids = [item.test_id for item in order_tests]
    mappings = list(
        db.scalars(
            select(AnalyzerTestMapping)
            .join(Analyzer, Analyzer.id == AnalyzerTestMapping.analyzer_id)
            .where(
                AnalyzerTestMapping.organization_id == context.organization_id,
                AnalyzerTestMapping.test_id.in_(test_ids),
                AnalyzerTestMapping.status == Status.active,
                Analyzer.organization_id == context.organization_id,
                Analyzer.branch_id == specimen.branch_id,
                Analyzer.status == Status.active,
            )
        ).all()
    )
    created = 0
    for mapping in mappings:
        existing = db.scalar(
            select(AnalyzerWorklistItem).where(
                AnalyzerWorklistItem.specimen_id == specimen.id,
                AnalyzerWorklistItem.analyzer_id == mapping.analyzer_id,
                AnalyzerWorklistItem.test_id == mapping.test_id,
            )
        )
        if existing:
            if existing.status == "cancelled":
                existing.status = "pending"
                existing.cancelled_reason = None
                existing.mapping_id = mapping.id
                existing.machine_test_code = mapping.machine_test_code
                existing.updated_by = context.user_id
                existing.correlation_id = request.state.correlation_id
                created += 1
            continue
        item = AnalyzerWorklistItem(
            organization_id=context.organization_id,
            branch_id=specimen.branch_id,
            specimen_id=specimen.id,
            order_id=specimen.order_id,
            test_id=mapping.test_id,
            analyzer_id=mapping.analyzer_id,
            mapping_id=mapping.id,
            machine_test_code=mapping.machine_test_code,
            status="pending",
            correlation_id=request.state.correlation_id,
            created_by=context.user_id,
            updated_by=context.user_id,
        )
        db.add(item)
        flush(db)
        record_event(
            db,
            request,
            context,
            event_type="analyzer.worklist_created",
            entity_type="analyzer_worklist_item",
            entity_id=item.id,
            branch_id=specimen.branch_id,
            action="create",
            new={
                "specimen_id": str(specimen.id),
                "analyzer_id": str(mapping.analyzer_id),
                "test_id": str(mapping.test_id),
                "machine_test_code": mapping.machine_test_code,
                "status": "pending",
            },
        )
        created += 1
    return created


def worklist_item_read(db: Session, item: AnalyzerWorklistItem) -> AnalyzerWorklistRead:
    specimen = db.get(Specimen, item.specimen_id)
    order = db.get(LabOrder, item.order_id)
    test = db.get(TestCatalogItem, item.test_id)
    analyzer = db.get(Analyzer, item.analyzer_id)
    latest_attempt = db.scalar(
        select(AnalyzerOrderAttempt)
        .where(AnalyzerOrderAttempt.worklist_item_id == item.id)
        .order_by(AnalyzerOrderAttempt.attempt_no.desc(), AnalyzerOrderAttempt.id.desc())
        .limit(1)
    )
    return AnalyzerWorklistRead(
        id=item.id,
        specimen_id=item.specimen_id,
        specimen_barcode=specimen.barcode if specimen else "Unknown",
        accession_number=specimen.accession_number if specimen else None,
        order_id=item.order_id,
        order_number=order.order_number if order else "Unknown",
        test_id=item.test_id,
        lis_test_code=test.code if test else "Unknown",
        test_name=test.name if test else "Unknown test",
        analyzer_id=item.analyzer_id,
        analyzer_code=analyzer.code if analyzer else "Unknown",
        analyzer_name=(f"{analyzer.vendor} {analyzer.model}" if analyzer else "Unknown analyzer"),
        mapping_id=item.mapping_id,
        machine_test_code=item.machine_test_code,
        status=item.status,
        correlation_id=item.correlation_id,
        cancelled_reason=item.cancelled_reason,
        latest_attempt_no=latest_attempt.attempt_no if latest_attempt else None,
        latest_attempt_state=latest_attempt.state if latest_attempt else None,
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


@router.get("/analyzer-worklist", response_model=Page[AnalyzerWorklistRead])
def list_analyzer_worklist(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
    status_filter: Annotated[str | None, Query(alias="status", max_length=30)] = None,
    analyzer_id: uuid.UUID | None = None,
) -> Page[AnalyzerWorklistRead]:
    filters = [AnalyzerWorklistItem.organization_id == context.organization_id]
    if not context.is_organization_scoped:
        filters.append(AnalyzerWorklistItem.branch_id.in_(context.branch_ids or {uuid.uuid4()}))
    if status_filter:
        filters.append(AnalyzerWorklistItem.status == status_filter.strip().lower())
    if analyzer_id:
        filters.append(AnalyzerWorklistItem.analyzer_id == analyzer_id)
    rows = db.execute(
        select(
            AnalyzerWorklistItem,
            func.count(AnalyzerWorklistItem.id).over().label("total_count"),
        )
        .where(*filters)
        .order_by(AnalyzerWorklistItem.created_at.desc(), AnalyzerWorklistItem.id.desc())
        .limit(limit)
        .offset(offset)
    ).all()
    total = rows[0].total_count if rows else 0
    if not rows and offset:
        total = db.scalar(select(func.count(AnalyzerWorklistItem.id)).where(*filters)) or 0
    items = [worklist_item_read(db, item) for item, _ in rows]
    return Page[AnalyzerWorklistRead](items=items, total=total, limit=limit, offset=offset)


@router.post(
    "/analyzer-worklist/{item_id}/enqueue",
    response_model=AnalyzerWorklistRead,
)
def enqueue_analyzer_worklist_item(
    item_id: uuid.UUID,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.manage"))],
) -> AnalyzerWorklistRead:
    item = db.scalar(
        select(AnalyzerWorklistItem).where(
            AnalyzerWorklistItem.id == item_id,
            AnalyzerWorklistItem.organization_id == context.organization_id,
        )
    )
    if item is None:
        raise HTTPException(status_code=404, detail="Worklist item not found")
    if not context.can_access_branch(item.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    if item.status not in {"pending", "failed"}:
        raise HTTPException(
            status_code=409,
            detail="Only pending or failed worklist items can be enqueued",
        )
    previous = {"status": item.status}
    item.status = "queued"
    item.updated_by = context.user_id
    attempt = create_queued_attempt(
        db,
        worklist_item=item,
        correlation_id=request.state.correlation_id,
        user_id=context.user_id,
    )
    record_event(
        db,
        request,
        context,
        event_type="analyzer.worklist_enqueued",
        entity_type="analyzer_worklist_item",
        entity_id=item.id,
        branch_id=item.branch_id,
        action="enqueue",
        previous=previous,
        new={"status": "queued", "attempt_id": str(attempt.id), "attempt_no": attempt.attempt_no},
    )
    commit(db)
    return worklist_item_read(db, item)


@router.post(
    "/analyzer-worklist/{item_id}/cancel",
    response_model=AnalyzerWorklistRead,
)
def cancel_analyzer_worklist_item(
    item_id: uuid.UUID,
    payload: AnalyzerWorklistCancel,
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.manage"))],
) -> AnalyzerWorklistRead:
    item = db.scalar(
        select(AnalyzerWorklistItem).where(
            AnalyzerWorklistItem.id == item_id,
            AnalyzerWorklistItem.organization_id == context.organization_id,
        )
    )
    if item is None:
        raise HTTPException(status_code=404, detail="Worklist item not found")
    if not context.can_access_branch(item.branch_id):
        raise HTTPException(status_code=403, detail="Branch access denied")
    if item.status in {"completed", "cancelled"}:
        raise HTTPException(status_code=409, detail="Worklist item is already closed")
    previous = {"status": item.status}
    item.status = "cancelled"
    item.cancelled_reason = payload.reason.strip() if payload.reason else None
    item.updated_by = context.user_id
    open_attempts = db.scalars(
        select(AnalyzerOrderAttempt).where(
            AnalyzerOrderAttempt.worklist_item_id == item.id,
            AnalyzerOrderAttempt.state.in_(("queued", "sending")),
        )
    ).all()
    for attempt in open_attempts:
        attempt.state = "failed"
        attempt.error = "Cancelled from worklist"
        attempt.finished_at = datetime.now(UTC)
    record_event(
        db,
        request,
        context,
        event_type="analyzer.worklist_cancelled",
        entity_type="analyzer_worklist_item",
        entity_id=item.id,
        branch_id=item.branch_id,
        action="cancel",
        previous=previous,
        new={"status": "cancelled", "reason": item.cancelled_reason},
    )
    commit(db)
    return worklist_item_read(db, item)


def order_attempt_read(attempt: AnalyzerOrderAttempt) -> AnalyzerOrderAttemptRead:
    return AnalyzerOrderAttemptRead(
        id=attempt.id,
        worklist_item_id=attempt.worklist_item_id,
        analyzer_id=attempt.analyzer_id,
        attempt_no=attempt.attempt_no,
        state=attempt.state,
        correlation_id=attempt.correlation_id,
        payload_hash=attempt.payload_hash,
        request_message_id=attempt.request_message_id,
        response_message_id=attempt.response_message_id,
        error=attempt.error,
        started_at=attempt.started_at,
        finished_at=attempt.finished_at,
        created_at=attempt.created_at,
        updated_at=attempt.updated_at,
    )


@router.get("/analyzer-orders/attempts", response_model=Page[AnalyzerOrderAttemptRead])
def list_analyzer_order_attempts(
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.read"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
    state: Annotated[str | None, Query(max_length=30)] = None,
    worklist_item_id: uuid.UUID | None = None,
) -> Page[AnalyzerOrderAttemptRead]:
    filters = [AnalyzerOrderAttempt.organization_id == context.organization_id]
    if not context.is_organization_scoped:
        filters.append(AnalyzerOrderAttempt.branch_id.in_(context.branch_ids or {uuid.uuid4()}))
    if state:
        filters.append(AnalyzerOrderAttempt.state == state.strip().lower())
    if worklist_item_id:
        filters.append(AnalyzerOrderAttempt.worklist_item_id == worklist_item_id)
    rows = db.execute(
        select(
            AnalyzerOrderAttempt,
            func.count(AnalyzerOrderAttempt.id).over().label("total_count"),
        )
        .where(*filters)
        .order_by(AnalyzerOrderAttempt.created_at.desc(), AnalyzerOrderAttempt.attempt_no.desc())
        .limit(limit)
        .offset(offset)
    ).all()
    total = rows[0].total_count if rows else 0
    if not rows and offset:
        total = db.scalar(select(func.count(AnalyzerOrderAttempt.id)).where(*filters)) or 0
    return Page[AnalyzerOrderAttemptRead](
        items=[order_attempt_read(item) for item, _ in rows],
        total=total,
        limit=limit,
        offset=offset,
    )


@router.post("/analyzer-orders/process", response_model=AnalyzerOrderProcessRead)
def process_analyzer_order_queue(
    request: Request,
    db: Db,
    context: Annotated[AuthContext, Depends(require_permission("analyzer.manage"))],
    limit: Annotated[int, Query(ge=1, le=100)] = 20,
) -> AnalyzerOrderProcessRead:
    attempts = process_queued_orders(db, request, context, limit=limit)
    commit(db)
    return AnalyzerOrderProcessRead(
        processed=len(attempts),
        attempts=[order_attempt_read(item) for item in attempts],
    )
